"""Write table data to XLSX with configurable column order."""

import re
import unicodedata
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Shared styling: all borders (thin), light gray fill, bold
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_LIGHT_GRAY_FILL = PatternFill(patternType="solid", fgColor="D3D3D3")
_BOLD_FONT = Font(bold=True)

from .config import OUTPUT_COLUMN_ORDER, OUTPUT_SHEET_NAME


def _normalize_header(name: str) -> str:
    """Normalize header for comparison: strip, collapse spaces (incl. newlines), uppercase, remove accents (e.g. CORREÇÃO → CORRECAO)."""
    s = re.sub(r"[\s\n\r]+", " ", str(name).strip()).upper()
    nfd = unicodedata.normalize("NFD", s)
    return "".join(c for c in nfd if unicodedata.category(c) != "Mn")


def _brazilian_to_numeric(ser: pd.Series) -> pd.Series:
    """
    Convert a pandas Series from Brazilian number strings to numeric (2 decimals).
    Brazilian: comma = decimal (reais, centavos), dot = thousands. Uses re for cleanup.
    """
    def one_cell(val):
        if val is None or (isinstance(val, str) and not str(val).strip()):
            return val
        if isinstance(val, (int, float)):
            return round(float(val), 2) if isinstance(val, float) else val
        s = re.sub(r"R\$\s*", "", str(val), flags=re.I)
        s = re.sub(r"[\s\u00a0]", "", s)
        s = s.strip()
        if not s:
            return val
        # Brazilian: comma = decimal, dot = thousands
        if "," in s:
            parts = re.split(r",", s, maxsplit=1)
            int_str = re.sub(r"\.", "", parts[0].strip())
            dec_str = (parts[1].strip() + "00")[:2]
            if re.match(r"^-?\d+$", int_str) and re.match(r"^\d{1,2}$", dec_str):
                try:
                    sign = -1 if int_str.startswith("-") else 1
                    return round(sign * (abs(int(int_str)) + int(dec_str) / 100.0), 2)
                except ValueError:
                    pass
        if "." in s and "," not in s:
            try:
                return int(re.sub(r"\.", "", s))
            except ValueError:
                pass
        if re.match(r"^-?\d+$", re.sub(r"\.", "", s)):
            try:
                return int(re.sub(r"\.", "", s))
            except ValueError:
                pass
        return val

    return ser.apply(one_cell)


def _convert_table_brazilian_pandas(
    table_part: list[list],
    venci_col_0based: int | None,
) -> list[list]:
    """
    Use pandas + re to convert table rows: Brazilian numbers → float (2 decimals).
    Returns list of rows (header + data) ready for Excel; VENCIMENTO column left as-is for date parsing.
    """
    if not table_part or len(table_part) < 2:
        return table_part
    header = table_part[0]
    df = pd.DataFrame(table_part[1:], columns=header)
    for col_idx, col_name in enumerate(header):
        if venci_col_0based is not None and col_idx == venci_col_0based:
            continue
        df[col_name] = _brazilian_to_numeric(df[col_name])
    # Back to list of rows; convert numpy types to Python for openpyxl
    out = [list(df.columns)]
    for _, row in df.iterrows():
        out.append([_pyval(x) for x in row.tolist()])
    return out


def _pyval(x) -> str | int | float | date:
    """Convert numpy/pandas value to Python type for openpyxl."""
    if pd.isna(x):
        return ""
    if hasattr(x, "item"):
        return x.item()
    return x


def _to_number(value: str) -> int | float | str:
    """
    Parse as Brazilian number: comma = decimal (reais, centavos), dot = thousands.
    E.g. 36,89 = 36 reais and 89 centavos → 36.89. Always returns at most 2 decimal places.
    """
    if value is None or (isinstance(value, str) and not value.strip()):
        return value if value is not None else ""
    s = str(value).strip()
    if not s:
        return s
    # Already numeric – round floats to 2 decimals
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return round(value, 2)
    # Strip R$ and spaces
    normalized = re.sub(r"^R\$\s*", "", s, flags=re.IGNORECASE)
    normalized = "".join(c for c in normalized if c not in " \u00a0\t").strip()
    if not normalized:
        return s
    # Keep only digits, one comma, dots, minus
    only_nums = "".join(c for c in normalized if c in "0123456789,.-")
    if not only_nums:
        return s
    # Brazilian rule: comma = decimal separator (centavos), dot = thousands
    if "," in only_nums:
        # One comma: part before = integer (remove dots), part after = decimals (max 2)
        parts = only_nums.split(",", 1)
        int_part = parts[0].replace(".", "").strip()
        dec_part = (parts[1].strip() + "00")[:2]  # at most 2 digits (centavos)
        if not int_part.lstrip("-").isdigit() or not dec_part.isdigit():
            return s
        try:
            sign = -1 if int_part.startswith("-") else 1
            num = abs(int(int_part)) + int(dec_part) / 100.0
            return round(sign * num, 2)
        except ValueError:
            return s
    # No comma: dot = thousands (e.g. 1.200) or plain integer
    no_dots = only_nums.replace(".", "")
    if no_dots.lstrip("-").isdigit():
        try:
            return int(no_dots)
        except ValueError:
            pass
    return s


def _parse_date(value: str) -> date | str:
    """Parse date string to datetime for Excel; Brazilian DD/MM/AAAA preferred. Returns original string if not parseable."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return value if value is not None else ""
    s = str(value).strip()
    if not s:
        return s
    if hasattr(value, "year"):  # already a date/datetime
        return value
    formats = ["%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%d-%m-%y"]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return s


def _autofit_columns(ws: Worksheet, padding: int = 2, max_width: int = 80) -> None:
    """Set column widths so content fits; avoid cut text. For wrap_text cells use max line length per column."""
    if ws.max_row == 0 or ws.max_column == 0:
        return
    for col_idx in range(1, ws.max_column + 1):
        width = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is not None:
                s = str(val).strip()
                if "\n" in s:
                    length = max(len(line) for line in s.splitlines()) if s else 0
                else:
                    length = len(s)
                if length > width:
                    width = length
        if width > 0:
            width = min(width + padding, max_width)
            ws.column_dimensions[get_column_letter(col_idx)].width = width


def reorder_columns(
    rows: list[list[str]],
    column_order: list[str] | list[int],
    has_header: bool = True,
) -> list[list[str]]:
    """
    Reorder columns so they match the desired order.

    If column_order contains strings, the first row is treated as header and
    column order is by header name. If column_order contains integers, they
    are 0-based column indices.

    Args:
        rows: Table rows (list of list of cell values).
        column_order: Desired column order (header names or indices).
        has_header: If True, first row is header and used for name-based order.

    Returns:
        New list of rows with columns reordered.
    """
    if not rows:
        return []

    if not column_order:
        return rows

    header = rows[0] if has_header else None
    data_rows = rows[1:] if has_header else rows

    # Build index mapping: output position -> source column index
    if column_order and isinstance(column_order[0], int):
        # Indices: assume column_order is list of 0-based indices
        index_map = list(column_order)
    else:
        # Header names: find index for each name (normalized match so PDF "Correção Monetária" matches config "CORREÇÃO MONETÁRIA")
        if not header:
            return rows
        normalized_to_index = {}
        for i, h in enumerate(header):
            key = _normalize_header(str(h).strip())
            if key not in normalized_to_index:
                normalized_to_index[key] = i
        index_map = []
        for name in column_order:
            norm_name = _normalize_header(str(name).strip())
            idx = normalized_to_index.get(norm_name)
            # Fallback: PDF may have "Correção" only (normalizes to "CORRECAO") for CORREÇÃO MONETÁRIA
            if idx is None and norm_name == "CORRECAO MONETARIA":
                idx = normalized_to_index.get("CORRECAO")
            if idx is not None:
                index_map.append(idx)

    def reorder_row(row: list[str]) -> list[str]:
        return [row[i] if i < len(row) else "" for i in index_map]

    out_header = reorder_row(header) if header else []
    out_data = [reorder_row(r) for r in data_rows]
    return [out_header] + out_data if out_header else out_data


def write_xlsx(
    rows: list[list[str]],
    output_path: str | Path,
    sheet_name: str | None = None,
    column_order: list[str] | list[int] | None = None,
    has_header: bool = True,
) -> Path:
    """
    Write rows to an XLSX file with optional column reordering.

    Args:
        rows: Table data (list of rows, each row is list of cells).
        output_path: Path for the output .xlsx file.
        sheet_name: Sheet name; uses config default if None.
        column_order: Desired column order; uses config if None.
        has_header: Whether first row is header (for column order by name).

    Returns:
        Resolved output path.
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    order = column_order if column_order is not None else OUTPUT_COLUMN_ORDER
    name = sheet_name if sheet_name is not None else OUTPUT_SHEET_NAME

    ordered = reorder_columns(rows, order, has_header=has_header) if order else rows

    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = name[:31]  # Excel sheet name length limit

    for row in ordered:
        ws.append(row)

    wb.save(path)
    return path.resolve()


def write_spreadsheet(
    header_rows: list[tuple[int, str, str]] | list[tuple[str, str]],
    table_rows: list[list[str]],
    output_path: str | Path,
    sheet_name: str | None = None,
    column_order: list[str] | list[int] | None = None,
    has_header: bool = True,
) -> Path:
    """
    Write a single sheet with (1) header section, (2) blank row, (3) table.

    Args:
        header_rows: Personal/header data. Either [(page_num, field, value), ...]
            (with page numbers from all pages) or [(field, value), ...] for a single block.
        table_rows: Table data (list of rows); first row is header if has_header.
        output_path: Path for the output .xlsx file.
        sheet_name: Sheet name; uses config default if None.
        column_order: Optional column order for the table section.
        has_header: Whether first row of table_rows is a header row.

    Returns:
        Resolved output path.
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    name = sheet_name if sheet_name is not None else OUTPUT_SHEET_NAME
    order = column_order if column_order is not None else OUTPUT_COLUMN_ORDER

    table_part = (
        reorder_columns(table_rows, order, has_header=has_header)
        if order and table_rows
        else table_rows
    )

    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = name[:31]

    # Header section: Field and Value only (no "Page" column, no "Field"/"Value" header row)
    header_section_start = 1
    if header_rows:
        if isinstance(header_rows[0][0], int):
            for row in header_rows:
                ws.append([row[1], row[2]])  # field, value (skip page)
        else:
            for row in header_rows:
                ws.append([row[0], row[1]])
        # Merge cells B to I on each header row (line level only)
        for row_idx in range(header_section_start, header_section_start + len(header_rows)):
            ws.merge_cells(
                start_row=row_idx,
                start_column=2,
                end_row=row_idx,
                end_column=9,
            )
    # Blank row to separate header from table (use [""] so openpyxl creates a real row)
    ws.append([""])

    # Style header section: all borders; column A bold + light gray
    if header_rows:
        num_header_rows = len(header_rows)
        for row_idx in range(header_section_start, header_section_start + num_header_rows):
            for col_idx in range(1, 10):  # A through I (merged B:I)
                c = ws.cell(row=row_idx, column=col_idx)
                c.border = _THIN_BORDER
                if col_idx == 1:
                    c.font = _BOLD_FONT
                    c.fill = _LIGHT_GRAY_FILL

    # Find VENCIMENTO column (0-based) for date formatting
    header_row = table_part[0] if table_part else []
    venci_col_0based = next(
        (i for i, h in enumerate(header_row) if str(h).strip().upper() == "VENCIMENTO"),
        None,
    )

    # Convert table with pandas + re: Brazilian numbers → float (2 decimals)
    table_part = _convert_table_brazilian_pandas(table_part, venci_col_0based)

    # Table: write row by row, cell by cell; numbers as numeric with exactly 2 decimals (Brazilian)
    table_start_row = ws.max_row + 1
    for row_offset, row in enumerate(table_part):
        row_idx = table_start_row + row_offset
        for col_idx, cell in enumerate(row, start=1):
            c = ws.cell(row=row_idx, column=col_idx)
            if row_offset == 0 and isinstance(cell, str) and "\n" in cell:
                c.value = cell
                c.alignment = Alignment(wrap_text=True)
            elif venci_col_0based is not None and (col_idx - 1) == venci_col_0based:
                val = _parse_date(cell)
                c.value = val
                if isinstance(val, date):
                    c.number_format = "dd/mm/yyyy"
            else:
                # Cell from pandas (float 2 decimals) or still string – ensure we write a number when possible
                if isinstance(cell, float):
                    val = round(cell, 2)
                elif isinstance(cell, int):
                    val = cell
                else:
                    val = _to_number(cell)  # fallback: parse Brazilian string (e.g. "260,000" → 260.0)
                c.value = val
                if isinstance(val, (int, float)):
                    # Force exactly 2 decimal places in display (Excel format 0.00)
                    c.number_format = "0.00"

    # Style table: all borders; first row (header) and last row (totals) bold + light gray
    num_table_cols = len(table_part[0]) if table_part else 0
    num_table_rows = len(table_part) if table_part else 0
    for row_offset in range(num_table_rows):
        row_idx = table_start_row + row_offset
        is_first_row = row_offset == 0
        is_last_row = num_table_rows > 1 and row_offset == num_table_rows - 1
        for col_idx in range(1, num_table_cols + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.border = _THIN_BORDER
            if is_first_row:
                c.font = _BOLD_FONT
                c.fill = _LIGHT_GRAY_FILL
            elif is_last_row:
                c.font = _BOLD_FONT
                c.fill = _LIGHT_GRAY_FILL
                if col_idx == 1:
                    c.value = "TOTAIS"

    # Second table: one blank row below main table, then fixed 5x2 table (columns A and B)
    _SECOND_TABLE_ROWS = [
        ("Correção monetária", "INPC"),
        ("Multa moratória", "2%"),
        ("Juros Moratórios (mensal)", "1%"),
        ("Result cobranças", "20%"),
        ("Honorários Advocatícios", "20%"),
    ]
    second_table_start_row = table_start_row + num_table_rows + 1
    for i, (label, value) in enumerate(_SECOND_TABLE_ROWS):
        row_idx = second_table_start_row + i
        for col_idx, val in enumerate((label, value), start=1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.value = val
            c.border = _THIN_BORDER
            if col_idx == 1:
                c.font = _BOLD_FONT

    _autofit_columns(ws)
    # Column B (VENCIMENTO): fit the header "VENCIMENTO" when bold
    ws.column_dimensions["B"].width = len("VENCIMENTO") + 3
    # Column H (HONORÁRIOS ADVOCATÍCIOS): slightly wider so "ADVOCATÍCIOS" stays on one line (avoid "S" on third line)
    hon_col_idx = next(
        (i + 1 for i, h in enumerate(header_row) if isinstance(h, str) and "ADVOCATÍCIOS" in h),
        8,
    )
    hon_letter = get_column_letter(hon_col_idx)
    ws.column_dimensions[hon_letter].width = max(
        ws.column_dimensions[hon_letter].width or 0,
        len("ADVOCATÍCIOS") + 4,
    )
    wb.save(path)
    return path.resolve()
